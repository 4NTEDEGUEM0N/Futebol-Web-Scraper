import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import load_workbook

google_colab = False
if google_colab:
  from google.colab import drive
  drive.mount('/content/drive')

times = {
    "São Paulo": "sao-paulo",
    "Botafogo": "botafogo",
    "Athletico PR": "athletico-pr",
    "CRB": "crb",
    "Flamengo": "flamengo",
    "Atlético GO": "atletico-go",
    "Corinthians": "corinthians",
    "Juventude": "juventude",
    "Vasco": "vasco",
    "RB Braga": "bragantino",
    "Atlético MG": "atletico-mg",
    "Bahia": "bahia",
    "Palmeiras": "palmeiras",
    "Grêmio": "gremio",
    "Fluminense": "fluminense",
    "Goiás": "goias",
    "San Lorenzo": "san-lorenzo",
    "Colo Colo": "colo-colo",
    "Peñarol": "penarol",
    "Talleres": "talleres-cordoba",
    "Nacional": "nacional",
    "Junior FC": "junior-barranquilla",
    "The Strongest": "the-strongest",
    "River Plate": "river-plate",
    "Bolívar": "bolivar"
}


def verificar_partidas_planilha(path_planilha):
    agora = datetime.now().strftime("%Y-%m-%d")
    wb = load_workbook(filename=path_planilha)
    sheet = wb['Palpi']
    partidas_atualizar = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not None:
            if isinstance(row[1], datetime):
                data_partida = row[1].strftime("%Y-%m-%d")
                if agora >= data_partida and row[4] is None:
                    time_casa = row[2]
                    time_fora = row[8]
                    placar = [row[4], row[6]]
                    partidas_atualizar.append([index, time_casa, time_fora, placar, data_partida])

    return partidas_atualizar, wb, sheet


def buscar_resultados(partidas_atualizar):
    partidas_resultado = []
    for partida_desatualizadas in partidas_atualizar:
        provavel_resultado = []
        url = "https://www.placardefutebol.com.br/time/" + str(times[partida_desatualizadas[1]]) + "/ultimos-jogos"
        r = requests.get(url)
        soup = BeautifulSoup(r.text, 'lxml')
        partidas = soup.find_all('div', attrs={'class': 'match__lg_card'})

        for chave, partida in enumerate(partidas):
            campeonato = partida.find('div', attrs={'class': 'match__lg_card--league'}).text
            if campeonato == "Copa do Brasil" or campeonato == "Copa Libertadores":
                if partida.find_all('div', attrs={'class': 'match__lg_card--date'}):
                    data = partida.find('div', attrs={'class': 'match__lg_card--date'}).text.split()
                    if data[0] == "ontem":
                        data_partida = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
                    else:
                        dia_mes = data[1].split("/")
                        data_partida = datetime(2024, int(dia_mes[1]), int(dia_mes[0])).strftime("%Y-%m-%d")

                    time_casa = partida.find('div', attrs={'class': 'match__lg_card--ht-name'}).text
                    time_fora = partida.find('div', attrs={'class': 'match__lg_card--at-name'}).text
                    placar = partida.find('div', attrs={'class': 'match__lg_card--scoreboard'}).text
                    placar = placar.split()
                    placar.pop(1)
                    provavel_resultado.append([campeonato, time_casa, time_fora, placar, data_partida])
        partidas_resultado.append(provavel_resultado)
    return partidas_resultado


def atualizar_partidas(partidas_atualizar, partidas_resultado, sheet):
    if not partidas_resultado:
        return "Nenhuma resultado encontrado"
    elif not partidas_atualizar:
        return "Nenhum jogo para atualizar"
    else:
        for chave, partida in enumerate(partidas_atualizar):
            for resultado in partidas_resultado[chave]:
                if resultado[4] == partida[4]:
                    sheet.cell(row=partida[0], column=5).value = float(resultado[3][0])
                    sheet.cell(row=partida[0], column=7).value = float(resultado[3][1])
    return sheet

if google_colab:
    path_planilha = "/content/drive/My Drive/Colab Notebooks/palpitometro.xlsx"
else:
    path_planilha = "/home/thiago/Área de trabalho/palpitometro.xlsx"

partidas_atualizar, wb, sheet = verificar_partidas_planilha(path_planilha)
partidas_resultado = buscar_resultados(partidas_atualizar)
sheet_atualizada = atualizar_partidas(partidas_atualizar, partidas_resultado, sheet)

for chave, partida in enumerate(partidas_atualizar):
    print(partida)
    print(partidas_resultado[chave])

for partida in partidas_atualizar:
    linha_valores = [cell.value for cell in sheet_atualizada[partida[0]]]
    print(linha_valores)

if google_colab:
    wb.save("/content/drive/My Drive/Colab Notebooks/alpitometro_atualizado.xlsx")
else:
    wb.save("/home/thiago/Área de trabalho/palpitometro_atualizado.xlsx")
